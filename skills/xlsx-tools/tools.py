"""
Excel Tools Skill

Provides Excel file manipulation tools using openpyxl and pandas.
"""

import logging
import os
from typing import Any, Dict, List, Optional, Union

from Jotty.core.infrastructure.utils.skill_status import SkillStatus
from Jotty.core.infrastructure.utils.tool_helpers import tool_error, tool_response, tool_wrapper

# Status emitter for progress updates
status = SkillStatus("xlsx-tools")


logger = logging.getLogger(__name__)


class ExcelToolkit:
    """Excel manipulation toolkit using openpyxl and pandas."""

    @staticmethod
    def _check_dependencies() -> Optional[Dict[str, Any]]:
        """Check if required dependencies are installed."""
        try:
            import openpyxl
            import pandas

            return None
        except ImportError as e:
            missing = str(e).split("'")[1] if "'" in str(e) else str(e)
            return {
                "success": False,
                "error": f"{missing} not installed. Install with: pip install openpyxl pandas",
            }

    @staticmethod
    def _validate_file_path(file_path: str, must_exist: bool = True) -> Optional[Dict[str, Any]]:
        """Validate file path."""
        if not file_path:
            return {"success": False, "error": "file_path parameter is required"}

        if must_exist and not os.path.exists(file_path):
            return {"success": False, "error": f"File not found: {file_path}"}

        if must_exist and not file_path.lower().endswith((".xlsx", ".xls", ".xlsm")):
            return {"success": False, "error": "File must be an Excel file (.xlsx, .xls, .xlsm)"}

        return None


@tool_wrapper()
def read_excel_tool(params: Dict[str, Any]) -> Dict[str, Any]:
    """
    Read Excel file to dictionary or dataframe representation.

    Args:
        params: Dictionary containing:
            - file_path (str, required): Path to the Excel file
            - sheet_name (str, optional): Sheet name to read (default: first sheet)
            - header_row (int, optional): Row number for column headers (0-indexed, default: 0)
            - as_dataframe (bool, optional): Return as dataframe dict (default: False, returns list of dicts)
            - skip_rows (int, optional): Number of rows to skip from top
            - max_rows (int, optional): Maximum number of rows to read

    Returns:
        Dictionary with:
            - success (bool): Whether operation succeeded
            - data (list/dict): Data from Excel file
            - columns (list): Column names
            - row_count (int): Number of rows
            - error (str, optional): Error message if failed
    """
    status.set_callback(params.pop("_status_callback", None))

    dep_error = ExcelToolkit._check_dependencies()
    if dep_error:
        return dep_error

    import pandas as pd

    file_path = params.get("file_path")
    path_error = ExcelToolkit._validate_file_path(file_path)
    if path_error:
        return path_error

    sheet_name = params.get("sheet_name", 0)  # Default to first sheet
    header_row = params.get("header_row", 0)
    as_dataframe = params.get("as_dataframe", False)
    skip_rows = params.get("skip_rows", None)
    max_rows = params.get("max_rows", None)

    try:
        df = pd.read_excel(
            file_path,
            sheet_name=sheet_name,
            header=header_row,
            skiprows=range(1, skip_rows + 1) if skip_rows else None,
            nrows=max_rows,
        )

        columns = df.columns.tolist()
        row_count = len(df)

        if as_dataframe:
            data = df.to_dict(orient="dict")
        else:
            data = df.to_dict(orient="records")

        logger.info(f"Read Excel file: {file_path}, sheet: {sheet_name}, rows: {row_count}")

        return {
            "success": True,
            "data": data,
            "columns": columns,
            "row_count": row_count,
            "sheet_name": sheet_name if isinstance(sheet_name, str) else f"Sheet{sheet_name + 1}",
        }

    except Exception as e:
        logger.error(f"Error reading Excel file: {e}", exc_info=True)
        return {"success": False, "error": f"Error reading Excel file: {str(e)}"}


@tool_wrapper()
def write_excel_tool(params: Dict[str, Any]) -> Dict[str, Any]:
    """
    Write data to Excel file.

    Args:
        params: Dictionary containing:
            - data (list/dict, required): Data to write (list of dicts or dataframe-style dict)
            - output_path (str, required): Output file path
            - sheet_name (str, optional): Sheet name (default: 'Sheet1')
            - index (bool, optional): Include index column (default: False)
            - overwrite (bool, optional): Overwrite existing file (default: True)

    Returns:
        Dictionary with:
            - success (bool): Whether operation succeeded
            - file_path (str): Path to written file
            - row_count (int): Number of rows written
            - error (str, optional): Error message if failed
    """
    status.set_callback(params.pop("_status_callback", None))

    dep_error = ExcelToolkit._check_dependencies()
    if dep_error:
        return dep_error

    import pandas as pd

    data = params.get("data")
    output_path = params.get("output_path")

    if data is None:
        return {"success": False, "error": "data parameter is required"}

    if not output_path:
        return {"success": False, "error": "output_path parameter is required"}

    sheet_name = params.get("sheet_name", "Sheet1")
    index = params.get("index", False)
    overwrite = params.get("overwrite", True)

    if not overwrite and os.path.exists(output_path):
        return {"success": False, "error": f"File already exists: {output_path}"}

    try:
        # Convert data to DataFrame
        if isinstance(data, list):
            df = pd.DataFrame(data)
        elif isinstance(data, dict):
            # Check if it's a dataframe-style dict (columns as keys with lists)
            if data and isinstance(list(data.values())[0], (list, dict)):
                df = pd.DataFrame(data)
            else:
                df = pd.DataFrame([data])
        else:
            return {"success": False, "error": "data must be a list of dicts or a dict"}

        # Ensure output directory exists
        output_dir = os.path.dirname(output_path)
        if output_dir:
            os.makedirs(output_dir, exist_ok=True)

        # Add .xlsx extension if not present
        if not output_path.lower().endswith(".xlsx"):
            output_path += ".xlsx"

        df.to_excel(output_path, sheet_name=sheet_name, index=index)

        logger.info(f"Wrote Excel file: {output_path}, rows: {len(df)}")

        return {
            "success": True,
            "file_path": output_path,
            "row_count": len(df),
            "sheet_name": sheet_name,
        }

    except Exception as e:
        logger.error(f"Error writing Excel file: {e}", exc_info=True)
        return {"success": False, "error": f"Error writing Excel file: {str(e)}"}


@tool_wrapper()
def get_sheets_tool(params: Dict[str, Any]) -> Dict[str, Any]:
    """
    List all sheets in an Excel workbook.

    Args:
        params: Dictionary containing:
            - file_path (str, required): Path to the Excel file

    Returns:
        Dictionary with:
            - success (bool): Whether operation succeeded
            - sheets (list): List of sheet names
            - count (int): Number of sheets
            - error (str, optional): Error message if failed
    """
    status.set_callback(params.pop("_status_callback", None))

    dep_error = ExcelToolkit._check_dependencies()
    if dep_error:
        return dep_error

    from openpyxl import load_workbook

    file_path = params.get("file_path")
    path_error = ExcelToolkit._validate_file_path(file_path)
    if path_error:
        return path_error

    try:
        wb = load_workbook(file_path, read_only=True)
        sheets = wb.sheetnames
        wb.close()

        logger.info(f"Listed sheets for: {file_path}, count: {len(sheets)}")

        return {"success": True, "sheets": sheets, "count": len(sheets)}

    except Exception as e:
        logger.error(f"Error getting sheets: {e}", exc_info=True)
        return {"success": False, "error": f"Error getting sheets: {str(e)}"}


@tool_wrapper()
def add_sheet_tool(params: Dict[str, Any]) -> Dict[str, Any]:
    """
    Add a new sheet to an existing Excel workbook.

    Args:
        params: Dictionary containing:
            - file_path (str, required): Path to the Excel file
            - sheet_name (str, required): Name for the new sheet
            - data (list, optional): Data to write (list of dicts)
            - position (int, optional): Position for new sheet (0-indexed, default: end)

    Returns:
        Dictionary with:
            - success (bool): Whether operation succeeded
            - sheet_name (str): Name of created sheet
            - file_path (str): Path to modified file
            - error (str, optional): Error message if failed
    """
    status.set_callback(params.pop("_status_callback", None))

    dep_error = ExcelToolkit._check_dependencies()
    if dep_error:
        return dep_error

    import pandas as pd
    from openpyxl import load_workbook

    file_path = params.get("file_path")
    sheet_name = params.get("sheet_name")
    data = params.get("data")
    position = params.get("position")

    path_error = ExcelToolkit._validate_file_path(file_path)
    if path_error:
        return path_error

    if not sheet_name:
        return {"success": False, "error": "sheet_name parameter is required"}

    try:
        wb = load_workbook(file_path)

        if sheet_name in wb.sheetnames:
            wb.close()
            return {"success": False, "error": f"Sheet already exists: {sheet_name}"}

        # Create new sheet
        if position is not None:
            ws = wb.create_sheet(sheet_name, position)
        else:
            ws = wb.create_sheet(sheet_name)

        # Add data if provided
        row_count = 0
        if data:
            df = pd.DataFrame(data) if isinstance(data, list) else pd.DataFrame([data])

            # Write headers
            for col_idx, col_name in enumerate(df.columns, 1):
                ws.cell(row=1, column=col_idx, value=col_name)

            # Write data
            for row_idx, row in enumerate(df.values, 2):
                for col_idx, value in enumerate(row, 1):
                    ws.cell(row=row_idx, column=col_idx, value=value)

            row_count = len(df)

        wb.save(file_path)
        wb.close()

        logger.info(f"Added sheet '{sheet_name}' to: {file_path}")

        return {
            "success": True,
            "sheet_name": sheet_name,
            "file_path": file_path,
            "row_count": row_count,
        }

    except Exception as e:
        logger.error(f"Error adding sheet: {e}", exc_info=True)
        return {"success": False, "error": f"Error adding sheet: {str(e)}"}


@tool_wrapper()
def update_cells_tool(params: Dict[str, Any]) -> Dict[str, Any]:
    """
    Update specific cells in an Excel workbook.

    Args:
        params: Dictionary containing:
            - file_path (str, required): Path to the Excel file
            - sheet_name (str, optional): Sheet name (default: first sheet)
            - updates (dict, required): Dictionary of cell:value pairs (e.g., {'A1': 'Hello', 'B2': 42})

    Returns:
        Dictionary with:
            - success (bool): Whether operation succeeded
            - updated_cells (int): Number of cells updated
            - file_path (str): Path to modified file
            - error (str, optional): Error message if failed
    """
    status.set_callback(params.pop("_status_callback", None))

    dep_error = ExcelToolkit._check_dependencies()
    if dep_error:
        return dep_error

    from openpyxl import load_workbook

    file_path = params.get("file_path")
    sheet_name = params.get("sheet_name")
    updates = params.get("updates")

    path_error = ExcelToolkit._validate_file_path(file_path)
    if path_error:
        return path_error

    if not updates or not isinstance(updates, dict):
        return {"success": False, "error": "updates parameter is required (dict of cell:value)"}

    try:
        wb = load_workbook(file_path)

        if sheet_name:
            if sheet_name not in wb.sheetnames:
                wb.close()
                return {"success": False, "error": f"Sheet not found: {sheet_name}"}
            ws = wb[sheet_name]
        else:
            ws = wb.active

        updated_count = 0
        for cell_ref, value in updates.items():
            ws[cell_ref] = value
            updated_count += 1

        wb.save(file_path)
        wb.close()

        logger.info(f"Updated {updated_count} cells in: {file_path}")

        return {
            "success": True,
            "updated_cells": updated_count,
            "file_path": file_path,
            "sheet_name": sheet_name or ws.title,
        }

    except Exception as e:
        logger.error(f"Error updating cells: {e}", exc_info=True)
        return {"success": False, "error": f"Error updating cells: {str(e)}"}


@tool_wrapper()
def add_formula_tool(params: Dict[str, Any]) -> Dict[str, Any]:
    """
    Add a formula to a cell in an Excel workbook.

    Args:
        params: Dictionary containing:
            - file_path (str, required): Path to the Excel file
            - sheet_name (str, optional): Sheet name (default: first sheet)
            - cell (str, required): Cell reference (e.g., 'C1')
            - formula (str, required): Excel formula (e.g., '=SUM(A1:B1)')

    Returns:
        Dictionary with:
            - success (bool): Whether operation succeeded
            - cell (str): Cell where formula was added
            - formula (str): Formula that was added
            - file_path (str): Path to modified file
            - error (str, optional): Error message if failed
    """
    status.set_callback(params.pop("_status_callback", None))

    dep_error = ExcelToolkit._check_dependencies()
    if dep_error:
        return dep_error

    from openpyxl import load_workbook

    file_path = params.get("file_path")
    sheet_name = params.get("sheet_name")
    cell = params.get("cell")
    formula = params.get("formula")

    path_error = ExcelToolkit._validate_file_path(file_path)
    if path_error:
        return path_error

    if not cell:
        return {"success": False, "error": "cell parameter is required"}

    if not formula:
        return {"success": False, "error": "formula parameter is required"}

    # Ensure formula starts with '='
    if not formula.startswith("="):
        formula = "=" + formula

    try:
        wb = load_workbook(file_path)

        if sheet_name:
            if sheet_name not in wb.sheetnames:
                wb.close()
                return {"success": False, "error": f"Sheet not found: {sheet_name}"}
            ws = wb[sheet_name]
        else:
            ws = wb.active

        ws[cell] = formula

        wb.save(file_path)
        wb.close()

        logger.info(f"Added formula to {cell} in: {file_path}")

        return {
            "success": True,
            "cell": cell,
            "formula": formula,
            "file_path": file_path,
            "sheet_name": sheet_name or ws.title,
        }

    except Exception as e:
        logger.error(f"Error adding formula: {e}", exc_info=True)
        return {"success": False, "error": f"Error adding formula: {str(e)}"}


@tool_wrapper()
def create_chart_tool(params: Dict[str, Any]) -> Dict[str, Any]:
    """
    Create a chart in an Excel workbook.

    Args:
        params: Dictionary containing:
            - file_path (str, required): Path to the Excel file
            - sheet_name (str, optional): Sheet name (default: first sheet)
            - chart_type (str, required): Type of chart ('bar', 'line', 'pie', 'area', 'scatter', 'column')
            - data_range (str, required): Data range for chart (e.g., 'A1:B10')
            - position (str, optional): Cell position for chart (default: 'E1')
            - title (str, optional): Chart title
            - categories_range (str, optional): Range for category labels (e.g., 'A2:A10')
            - width (int, optional): Chart width in EMUs (default: 15)
            - height (int, optional): Chart height in EMUs (default: 10)

    Returns:
        Dictionary with:
            - success (bool): Whether operation succeeded
            - chart_type (str): Type of chart created
            - position (str): Cell position of chart
            - file_path (str): Path to modified file
            - error (str, optional): Error message if failed
    """
    status.set_callback(params.pop("_status_callback", None))

    dep_error = ExcelToolkit._check_dependencies()
    if dep_error:
        return dep_error

    from openpyxl import load_workbook
    from openpyxl.chart import AreaChart, BarChart, LineChart, PieChart, Reference, ScatterChart

    file_path = params.get("file_path")
    sheet_name = params.get("sheet_name")
    chart_type = params.get("chart_type", "").lower()
    data_range = params.get("data_range")
    position = params.get("position", "E1")
    title = params.get("title", "")
    categories_range = params.get("categories_range")
    width = params.get("width", 15)
    height = params.get("height", 10)

    path_error = ExcelToolkit._validate_file_path(file_path)
    if path_error:
        return path_error

    if not chart_type:
        return {"success": False, "error": "chart_type parameter is required"}

    if not data_range:
        return {"success": False, "error": "data_range parameter is required"}

    # Chart type mapping
    chart_classes = {
        "bar": BarChart,
        "column": BarChart,
        "line": LineChart,
        "pie": PieChart,
        "area": AreaChart,
        "scatter": ScatterChart,
    }

    if chart_type not in chart_classes:
        return {
            "success": False,
            "error": f'Invalid chart_type. Supported types: {", ".join(chart_classes.keys())}',
        }

    try:
        wb = load_workbook(file_path)

        if sheet_name:
            if sheet_name not in wb.sheetnames:
                wb.close()
                return {"success": False, "error": f"Sheet not found: {sheet_name}"}
            ws = wb[sheet_name]
        else:
            ws = wb.active

        # Parse data range
        from openpyxl.utils import range_boundaries

        min_col, min_row, max_col, max_row = range_boundaries(data_range)

        # Create chart
        ChartClass = chart_classes[chart_type]
        chart = ChartClass()

        if title:
            chart.title = title

        chart.width = width
        chart.height = height

        # For column charts (vertical bars)
        if chart_type == "column":
            chart.type = "col"

        # Create data reference
        data = Reference(ws, min_col=min_col, min_row=min_row, max_col=max_col, max_row=max_row)

        # Add data to chart
        if chart_type == "scatter":
            # Scatter charts need x and y values
            chart.add_data(data, titles_from_data=True)
        else:
            chart.add_data(data, titles_from_data=True)

        # Add categories if provided
        if categories_range:
            cat_min_col, cat_min_row, cat_max_col, cat_max_row = range_boundaries(categories_range)
            categories = Reference(
                ws,
                min_col=cat_min_col,
                min_row=cat_min_row,
                max_col=cat_max_col,
                max_row=cat_max_row,
            )
            chart.set_categories(categories)

        # Add chart to worksheet
        ws.add_chart(chart, position)

        wb.save(file_path)
        wb.close()

        logger.info(f"Created {chart_type} chart at {position} in: {file_path}")

        return {
            "success": True,
            "chart_type": chart_type,
            "position": position,
            "data_range": data_range,
            "file_path": file_path,
            "sheet_name": sheet_name or ws.title,
        }

    except Exception as e:
        logger.error(f"Error creating chart: {e}", exc_info=True)
        return {"success": False, "error": f"Error creating chart: {str(e)}"}
